from openpyxl import load_workbook, Workbook

wb = load_workbook('userDB.xlsx')
ws = wb.active

def adduser(name,num):
    for r in range(2,ws.max_row+2):
        if ws.cell(r,1).value == name:
            return False
        if ws.cell(r,1).value is None:
            ws.cell(row=r,column=1,value=name)
            ws.cell(row=r,column=2,value=num)
            wb.save('userDB.xlsx')
            return True

def userlist():
    _list = []
    for r in range(2,ws.max_row+1):
        if ws.cell(r,1).value == None:
            break
        _list.append(ws.cell(r,1).value)
    return _list

def user_num(name):
    for r in range(2,ws.max_row+2):
        if ws.cell(r,1).value == name:
            return ws.cell(r,2).value

def change_num(name,num):
    for r in range(2,ws.max_row+2):
        if ws.cell(r,1).value == name:
            ws.cell(row=r,column=2,value=num)
            wb.save('userDB.xlsx')

def numreset():
    for r in range(2,ws.max_row+1):
        ws.cell(row=r,column=2,value=1)
    wb.save('userDB.xlsx')     


# numreset()
